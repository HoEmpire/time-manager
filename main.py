'''
test case
11.22-13.22 魔域 3
23.10-1.20 打游戏 6
23.10-1.20. 打游戏 6
'''
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

all_tpye = ['生活必须', '干活', '专业发展', '个人发展', '运动', '享受', '浪费']


def deleterows(worksheet, row_num):
    for i in range(worksheet.max_col):
        worksheet[i][row_num].value = None


class time_24h():
    def __init__(self, time):
        comma_index = time.find('.')
        self.hour = int(time[:comma_index])
        self.minute = int(time[comma_index + 1:len(time)])
        #print(self.hour)
        #print(self.minute)

    def __sub__(self, other):
        duration = 60 * (self.hour - other.hour) + (self.minute - other.minute)
        if duration < 0:
            duration += 24 * 60
        return duration


class event():
    def __init__(self, block_event_list, last_time):
        self.current_time = block_event_list[0]
        self.last_time = last_time
        self.content = block_event_list[1]
        self.label = int(block_event_list[2])
        self.type = all_tpye[int(self.label - 1)]
        self.cal_duration()

    def cal_duration(self):
        self.time_start = time_24h(self.last_time)
        self.time_end = time_24h(self.current_time)
        self.duration = self.time_end - self.time_start
        #print(self.duration)

    def report(self):
        print('\n*****记录成功*****')
        print('开始时间:{}时{}分'.format(self.time_start.hour,
                                   self.time_start.minute))
        print('结束时间:{}时{}分'.format(self.time_end.hour, self.time_end.minute))
        print('持续时间:{}分钟'.format(self.duration))
        print('内容:{}'.format(self.content))
        print('类型:{}'.format(self.type))
        print('****************')


class workbook():
    def __init__(self, adr):

        print('*****打开表格表*****')
        self.wb = load_workbook(adr, keep_vba=True)
        self.adr = adr
        self.ws = self.wb[self.wb.sheetnames[-1]]

        print('今天是{}，今天也要肝爆'.format(self.wb.sheetnames[-1]))
        self.find_end()
        self.get_last_time()

    def find_end(self):
        for i in range(2, self.ws.max_row):
            if (self.ws[i][0].value is None) or (self.ws[i][0].value == 0):
                self.end_row = i
                break
            #else:
            #print(self.ws[i][0].value)

    def get_last_time(self):
        if self.end_row == 2:
            self.last_time = input("设置元气一天的开始锚点(xx.xx)")
            #print(self.last_time)
        else:
            time = self.ws[self.end_row - 1][0].value
            index_minus = time.find('-')
            self.last_time = time[index_minus + 1:len(time)]
            #print(self.last_time)

    def add_new_event(self, new_event_block):
        if self.input_event_verify(new_event_block):
            self.new_event = event(new_event_block, self.last_time)
            self.new_event.report()
            self.ws[self.end_row][
                0].value = self.new_event.last_time + '-' + self.new_event.current_time
            self.ws[self.end_row][1].value = self.new_event.content
            self.ws[self.end_row][self.new_event.label +
                                  1].value = self.new_event.duration
            self.end_row += 1
        else:
            print('小伙子输入格式错误')

    def cal_sum(self):
        for j in range(2, 9):
            self.ws[self.ws.max_row - 2][j].value = 0
            for i in range(2, self.end_row):
                if self.ws[i][j].value is not None:
                    if self.ws[self.ws.max_row - 2][j].value is None:
                        self.ws[self.ws.max_row - 2][j].value = int(
                            self.ws[i][j].value)
                    else:
                        self.ws[self.ws.max_row - 2][j].value += int(
                            self.ws[i][j].value)

    def save_table(self):
        self.cal_sum()
        self.wb.save(self.adr)

    def input_event_verify(self, block_event_list):
        is_input_right = True
        if (block_event_list[0].count('.') != 1) or (int(
                block_event_list[2]) < 1) or (int(block_event_list[2]) > 7):
            is_input_right = False
        self.find_end()
        self.get_last_time()
        return is_input_right

    def add_new_sheet(self):
        print('*****创建新表*****')
        sheet_title = input('今天几号？')
        self.ws = self.wb.copy_worksheet(self.wb.worksheets[0])
        self.ws.title = sheet_title
        self.find_end()
        self.get_last_time()
        print('*****新表创建完成*****')

    def day_report(self):
        self.cal_sum()
        print('**********今天情况***********')
        sum_effective_time = 0
        for i in range(0, 7):
            sumtime = self.ws[self.ws.max_row - 2][i + 2].value
            print("{:8}\t{:3}小时{}分钟".format(all_tpye[i], int(sumtime / 60),
                                            sumtime % 60))
            if 0 < i < 5:
                sum_effective_time += sumtime
        print("{:8}\t{:3}小时{}分钟".format(
            '有效时间', int(sum_effective_time / 60), sum_effective_time % 60))
        print('***************************')

    def day_report_detail(self):

        print(
            '**************************************************************今天详细情况***************************************************************'
        )
        for i in range(1, self.end_row):
            time = self.ws[i][0].value
            content = self.ws[i][1].value

            type = []
            for j in range(2, 9):
                if self.ws[i][j].value is None:
                    type.append('0')
                else:
                    type.append(self.ws[i][j].value)
            print(
                '{:12}\t{:8}\t{:^8}\t{:^8}\t{:^8}\t{:^8}\t{:^8}\t{:^8}\t{:^8}'.
                format(time, content, type[0], type[1], type[2], type[3],
                       type[4], type[5], type[6]))
        self.get_last_time()
        print(
            '*****************************************************************************************************************************************'
        )

    def set_format(self):
        for row in self.ws.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')

    def delete_last_one(self):
        for i in range(0,8):
            self.ws[self.end_row-1][i].value = None
        self.end_row -= 1
        self.find_end()
        self.get_last_time()
        print('\n*****删除完成*****\n')


def input_new_event_block():
    onscreen = "输入事件(xx.xx 内容 分类):\n"
    onscreen += "1-生活必须\n"
    onscreen += "2-干活\n"
    onscreen += "3-专业发展\n"
    onscreen += "4-个人发展\n"
    onscreen += "5-运动\n"
    onscreen += "6-享受\n"
    onscreen += "7-浪费\n\n"

    str = input(onscreen)
    block_event_list = []

    if str.count(' ') != 2:
        print('小伙子输入的长度错误')
        return ' '
    else:
        space_index1 = str.find(' ')
        space_index2 = str.find(' ', space_index1 + 1)
        block_event_list.append(str[:space_index1])
        block_event_list.append(str[space_index1 + 1:space_index2])
        block_event_list.append(str[space_index2 + 1:])
        #print(block_event_list)
    return block_event_list


def main():
    adr = 'D:\时间管理.xlsx'
    mytable = workbook(adr)

    while True:
        task = input(
            "\n要干啥？\n(回车-添加事件\n1-创建新的一天\n2-当日时间安排汇报\n3-详细汇报\nd-删除最后一项\q-退出)\n")
        if task == '':
            event_block = input_new_event_block()
            if len(event_block) > 2:
                mytable.add_new_event(event_block)
                mytable.save_table()
        elif task == '1':
            mytable.add_new_sheet()
            mytable.save_table()
        elif task == '2':
            mytable.day_report()
        elif task == '3':
            mytable.day_report_detail()
        elif task == 'd':
            mytable.delete_last_one()
            mytable.save_table()
        elif task == 'q':
            break

    mytable.set_format()
    mytable.save_table()


if __name__ == '__main__':
    main()
#input_new_event()
